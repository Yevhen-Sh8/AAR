from datetime import date

from httpx import ASGITransport, AsyncClient
from sqlalchemy.ext.asyncio import async_sessionmaker

from aar_api.core.db import _engine
from aar_api.main import app
from aar_api.models.dictionaries import ItemType, LossReason, Operator, Zone
from aar_api.models.event import Item, Outcome, UsageEvent


async def _seed_two_months() -> None:
    Session = async_sessionmaker(_engine, expire_on_commit=False)
    async with Session() as s:
        a = ItemType(code="A", name_uk="A")
        e1 = Operator(code="E-01", name_uk="01")
        e2 = Operator(code="E-02", name_uk="02")
        lr_op = LossReason(code="op", name_uk="op", zone=Zone.OPERATOR)
        lr_ext = LossReason(code="ext", name_uk="ext", zone=Zone.EXTERNAL)
        s.add_all([a, e1, e2, lr_op, lr_ext])
        await s.flush()
        items = [Item(serial_no=f"A-{i:05d}", item_type_id=a.id) for i in range(20)]
        s.add_all(items)
        await s.flush()

        def ev(i: int, op_id: int, d: date, **kw: object) -> UsageEvent:
            return UsageEvent(item_id=items[i].id, operator_id=op_id,
                              event_date=d, **kw)  # type: ignore[arg-type]

        nov = date(2025, 11, 15)
        dec = date(2025, 12, 15)
        # November E-01: 8 launches, 6 success, 1 lost(external), 1 lost(operator)
        s.add_all([
            ev(0, e1.id, nov, outcome=Outcome.SUCCESS),
            ev(1, e1.id, nov, outcome=Outcome.SUCCESS),
            ev(2, e1.id, nov, outcome=Outcome.SUCCESS),
            ev(3, e1.id, nov, outcome=Outcome.SUCCESS),
            ev(4, e1.id, nov, outcome=Outcome.SUCCESS),
            ev(5, e1.id, nov, outcome=Outcome.SUCCESS),
            ev(6, e1.id, nov, outcome=Outcome.LOST, loss_reason_id=lr_ext.id),
            ev(7, e1.id, nov, outcome=Outcome.LOST, loss_reason_id=lr_op.id),
            # E-02 Nov: 4 launches, 4 success
            ev(8, e2.id, nov, outcome=Outcome.SUCCESS),
            ev(9, e2.id, nov, outcome=Outcome.SUCCESS),
            ev(10, e2.id, nov, outcome=Outcome.SUCCESS),
            ev(11, e2.id, nov, outcome=Outcome.SUCCESS),
            # December E-01: 5 launches, 5 success (improvement)
            ev(12, e1.id, dec, outcome=Outcome.SUCCESS),
            ev(13, e1.id, dec, outcome=Outcome.SUCCESS),
            ev(14, e1.id, dec, outcome=Outcome.SUCCESS),
            ev(15, e1.id, dec, outcome=Outcome.SUCCESS),
            ev(16, e1.id, dec, outcome=Outcome.SUCCESS),
        ])
        await s.commit()


async def test_monthly_msr_c_and_trend() -> None:
    await _seed_two_months()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        r = await client.get("/reports/monthly", params={"year": 2025, "month": 12})
        assert r.status_code == 200, r.text
        body = r.json()
        rows = {row["operator_code"]: row for row in body["rows"]}
        assert rows["E-01"]["launched"] == 5
        assert rows["E-01"]["success"] == 5
        assert rows["E-01"]["msr"] == 1.0
        # Dec MSR=1.0 vs Nov MSR=6/8=0.75 → Δη=+25.0 pp
        assert abs(rows["E-01"]["delta_msr_pp"] - 25.0) < 0.01

        rating = {r["operator_code"]: r for r in body["rating"]}
        # ADR-027: five sorties cannot carry a readiness verdict, however good
        # the arithmetic looks. This assertion used to read `== "high"`, which
        # is exactly the over-claim the floor exists to stop.
        assert rating["E-01"]["category"] == "insufficient_data"
        assert rating["E-01"]["sample_sufficient"] is False
        assert rating["E-01"]["sorties"] == 5
        assert rating["E-01"]["msr_c"] == 1.0
        assert rating["E-01"]["rank"] == 1

        trends = {t["operator_code"]: t for t in body["trends"]}
        assert trends["E-01"]["trend"] == "up"
        assert trends["E-01"]["msr_prev"] == 0.75


async def test_monthly_zones_aggregation_november() -> None:
    await _seed_two_months()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        r = await client.get("/reports/monthly", params={"year": 2025, "month": 11})
        body = r.json()
        zones = {z["zone"]: z for z in body["zones"]}
        assert zones["operator"]["losses"] == 1
        assert zones["external"]["losses"] == 1
        # η_c for E-01 in Nov = 6 / (8 - 1) = 6/7 ≈ 0.8571
        rows = {row["operator_code"]: row for row in body["rows"]}
        assert abs(rows["E-01"]["msr_c"] - round(6 / 7, 4)) < 1e-4


async def test_monthly_xlsx_pdf() -> None:
    await _seed_two_months()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        r = await client.get("/reports/monthly.xlsx", params={"year": 2025, "month": 12})
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
        r = await client.get("/reports/monthly.pdf", params={"year": 2025, "month": 12})
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"


async def test_monthly_xlsx_labels_are_readable_without_the_notation_doc() -> None:
    """The printed report has to explain itself.

    Column heads used to be bare `η_c (MSR_c)` — a symbol that appears nowhere
    else in the product and was never defined on any screen. A monthly report
    goes to a commander who has not read `docs/metrics.md`, so the head now
    carries the Ukrainian name (notation kept in brackets for traceability) and
    a «Позначення» block spells out every formula under the tables.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    await _seed_two_months()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        r = await client.get("/reports/monthly.xlsx", params={"year": 2025, "month": 12})
        assert r.status_code == 200

    ws = load_workbook(BytesIO(r.content)).active
    assert ws is not None
    text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )

    # Plain name first, notation only as a bracketed hint.
    assert "Успішність, % (η)" in text
    assert "Успішність обслуги, % (η_c)" in text
    assert "Втрати обслуги, % (λ_c)" in text
    assert "Зміна, в.п. (Δη)" in text

    # And the legend that makes the sheet self-contained.
    assert "Позначення" in text
    assert "успішні ÷ запущені" in text
    assert "85%" in text  # rating thresholds spelled out




async def test_a_verdict_appears_once_there_are_enough_sorties() -> None:
    """The floor withholds the category, it does not withhold the number."""
    Session = async_sessionmaker(_engine, expire_on_commit=False)
    async with Session() as s:
        a = ItemType(code="A", name_uk="A")
        op = Operator(code="E-09", name_uk="09")
        s.add_all([a, op])
        await s.flush()
        items = [Item(serial_no=f"B-{i:05d}", item_type_id=a.id) for i in range(12)]
        s.add_all(items)
        await s.flush()
        for i in range(12):
            s.add(
                UsageEvent(
                    item_id=items[i].id,
                    operator_id=op.id,
                    event_date=date(2025, 12, 15),
                    outcome=Outcome.SUCCESS,
                )
            )
        await s.commit()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        body = (
            await client.get("/reports/monthly", params={"year": 2025, "month": 12})
        ).json()

    row = {r["operator_code"]: r for r in body["rating"]}["E-09"]
    assert row["sorties"] == 12
    assert row["sample_sufficient"] is True
    assert row["category"] == "high"


async def test_a_thin_sample_never_tops_the_table() -> None:
    """A 100% built on two flights must not outrank a measured operator."""
    Session = async_sessionmaker(_engine, expire_on_commit=False)
    async with Session() as s:
        a = ItemType(code="A", name_uk="A")
        thin = Operator(code="E-THIN", name_uk="thin")
        solid = Operator(code="E-SOLID", name_uk="solid")
        s.add_all([a, thin, solid])
        await s.flush()
        items = [Item(serial_no=f"C-{i:05d}", item_type_id=a.id) for i in range(14)]
        s.add_all(items)
        await s.flush()
        # Two perfect flights…
        for i in range(2):
            s.add(UsageEvent(item_id=items[i].id, operator_id=thin.id,
                             event_date=date(2025, 12, 15), outcome=Outcome.SUCCESS))
        # …against twelve, one of which failed.
        for i in range(2, 13):
            s.add(UsageEvent(item_id=items[i].id, operator_id=solid.id,
                             event_date=date(2025, 12, 15), outcome=Outcome.SUCCESS))
        s.add(UsageEvent(item_id=items[13].id, operator_id=solid.id,
                         event_date=date(2025, 12, 15), outcome=Outcome.REPAIR))
        await s.commit()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        body = (
            await client.get("/reports/monthly", params={"year": 2025, "month": 12})
        ).json()

    ranks = {r["operator_code"]: r["rank"] for r in body["rating"]}
    assert ranks["E-SOLID"] < ranks["E-THIN"]
