from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from aar_api.schemas.monthly import MonthlyReport
from aar_api.schemas.reports import DailyReport

# Друкований звіт іде людині, яка не читала `docs/metrics.md`. Тому в
# заголовках — звичайна українська назва, а наукова нотація (η, η_c, λ_c)
# лишається у дужках і в легенді: за нею показник можна звірити з літературою
# і з полями API (`msr`, `msr_c`, `clr`). Легенда друкується під таблицею,
# щоб звіт був самодостатнім без інших документів.
H_MSR = "Успішність, % (η)"
H_MSR_C = "Успішність обслуги, % (η_c)"
H_CLR = "Втрати обслуги, % (λ_c)"
H_DELTA = "Зміна, в.п. (Δη)"

LEGEND_MSR = "Успішність (η, MSR) = успішні ÷ запущені; зриви до запуску в знаменник не входять."
LEGEND_MSR_C = (
    "Успішність обслуги (η_c, MSR_c) = успішні ÷ (запущені − втрати з зовнішніх "
    "причин − втрати через заводський дефект). Пороги: ≥ 85% висока готовність, "
    "70–85% задовільна, < 70% потребує до-підготовки."
)
LEGEND_CLR = (
    "Втрати обслуги (λ_c, CLR) = втрати й ремонти із зоною «обслуга» ÷ запущені."
)
LEGEND_DELTA = "Зміна (Δη) — різниця успішності з попереднім періодом, у відсоткових пунктах."
LEGEND_SAMPLE = (
    "«Запусків» — скільки застосувань стоїть за числом. Менш ніж 10 запусків не "
    "дають підстав для висновку про готовність: категорія тоді — «замало даних»."
)

#: Readiness buckets in Ukrainian. The raw enum used to be printed verbatim
#: into a document going to a commander.
CATEGORY_UK = {
    "high": "висока готовність",
    "ok": "задовільна",
    "needs_training": "потребує до-підготовки",
    "insufficient_data": "замало даних для висновку",
}


def daily_report_to_xlsx(report: DailyReport) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = f"Доба {report.report_date.isoformat()}"
    bold = Font(bold=True)

    ws.append([f"Аналітична довідка за {report.report_date.isoformat()}"])
    ws["A1"].font = bold
    ws.append([])
    ws.append(["Т.1. Зведені показники"])
    ws.cell(row=ws.max_row, column=1).font = bold
    headers = ["Експлуатант", "Тип", "Запущено", "Втрачено", "Ремонт", "Успіх", H_MSR]
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.font = bold
    for r in report.rows:
        ws.append([r.operator_code, r.item_type_code, r.launched, r.lost,
                   r.repaired, r.success, r.msr])
    t = report.totals
    ws.append([t.operator_code, t.item_type_code, t.launched, t.lost,
               t.repaired, t.success, t.msr])
    for c in ws[ws.max_row]:
        c.font = bold

    ws.append([])
    ws.append(["Позначення"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([LEGEND_MSR])

    ws.append([])
    ws.append(["Т.2. Безповоротні втрати — деталізація"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Експлуатант", "Тип", "Серійний №", "Причина", "Примітка"])
    for c in ws[ws.max_row]:
        c.font = bold
    for d in report.loss_details:
        ws.append([d.operator_code, d.item_type_code, d.serial_no, d.reason_code, d.notes or ""])

    ws.append([])
    ws.append(["Т.2.1. Розподіл втрат за причинами"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Тип", "Причина", "Зона", "Кількість"])
    for c in ws[ws.max_row]:
        c.font = bold
    for b in report.loss_breakdown:
        ws.append([b.item_type_code, b.reason_code, b.zone.value, b.count])

    ws.append([])
    ws.append(["Т.3. Повернення в ремонт — деталізація"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Експлуатант", "Тип", "Серійний №", "Причина", "Примітка"])
    for c in ws[ws.max_row]:
        c.font = bold
    for rd in report.repair_details:
        ws.append(
            [rd.operator_code, rd.item_type_code, rd.serial_no, rd.reason_code, rd.notes or ""]
        )

    ws.append([])
    ws.append(["Т.3.1. Розподіл повернень за причинами"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Тип", "Причина", "Зона", "Кількість"])
    for c in ws[ws.max_row]:
        c.font = bold
    for rb in report.repair_breakdown:
        ws.append([rb.item_type_code, rb.reason_code, rb.zone.value, rb.count])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _table(data: list[list[object]]) -> Table:
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a3a1a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    return tbl


def daily_report_to_pdf(report: DailyReport) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Daily AAR report")
    styles = getSampleStyleSheet()
    story: list = [
        Paragraph(f"Аналітична довідка за {report.report_date.isoformat()}", styles["Title"]),
        Spacer(1, 12),
        Paragraph("Т.1. Зведені показники", styles["Heading2"]),
    ]
    summary: list[list[object]] = [
        ["Експлуатант", "Тип", "Запущ.", "Втрач.", "Ремонт", "Успіх", H_MSR]
    ]
    for r in report.rows:
        summary.append(
            [r.operator_code, r.item_type_code, r.launched, r.lost,
             r.repaired, r.success, f"{r.msr:.2%}"]
        )
    t = report.totals
    summary.append(
        [t.operator_code, t.item_type_code, t.launched, t.lost,
         t.repaired, t.success, f"{t.msr:.2%}"]
    )
    story.append(_table(summary))
    story.append(Spacer(1, 6))
    story.append(Paragraph(LEGEND_MSR, styles["BodyText"]))

    def _break_table(title: str, items: list) -> None:
        if not items:
            return
        story.extend([Spacer(1, 12), Paragraph(title, styles["Heading2"])])
        rows: list[list[object]] = [["Тип", "Причина", "Зона", "Кількість"]]
        rows.extend(
            [b.item_type_code, b.reason_code, b.zone.value, b.count] for b in items
        )
        story.append(_table(rows))

    _break_table("Т.2.1. Розподіл втрат", report.loss_breakdown)
    _break_table("Т.3.1. Розподіл повернень", report.repair_breakdown)

    doc.build(story)
    return buf.getvalue()


def monthly_report_to_xlsx(report: MonthlyReport) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = f"{report.year}-{report.month:02d}"
    bold = Font(bold=True)

    ws.append([f"Звіт за {report.year}-{report.month:02d}"])
    ws["A1"].font = bold
    ws.append([])
    ws.append(["Т.4. Інтегральні показники по експлуатантах"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([
        "Експлуатант", "Тип", "Запущ.", "Втрач.", "Ремонт", "Успіх",
        H_MSR, H_MSR_C, H_CLR, H_DELTA,
    ])
    for c in ws[ws.max_row]:
        c.font = bold
    for r in report.rows:
        ws.append([
            r.operator_code, r.item_type_code, r.launched, r.lost, r.repaired,
            r.success, r.msr, r.msr_c, r.clr, r.delta_msr_pp,
        ])
    t = report.totals
    ws.append([
        t.operator_code, t.item_type_code, t.launched, t.lost, t.repaired,
        t.success, t.msr, t.msr_c, t.clr, t.delta_msr_pp,
    ])
    for c in ws[ws.max_row]:
        c.font = bold

    ws.append([])
    ws.append(["Рейтинг експлуатантів за успішністю обслуги (η_c)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Місце", "Експлуатант", H_MSR_C, "Запусків", "Категорія"])
    for c in ws[ws.max_row]:
        c.font = bold
    for rt in report.rating:
        ws.append([rt.rank, rt.operator_code, rt.msr_c, rt.sorties, CATEGORY_UK[rt.category]])

    ws.append([])
    ws.append(["Т.7. Зони відповідальності"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Зона", "Втрати", "Ремонти", "Разом", "Частка від запущених"])
    for c in ws[ws.max_row]:
        c.font = bold
    for z in report.zones:
        ws.append([z.zone.value, z.losses, z.repairs, z.total, z.share_of_launched])

    ws.append([])
    ws.append(["Т.6. Динаміка vs попередній місяць"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Експлуатант", "Успішність попер.", "Успішність поточ.", "Тренд"])
    for c in ws[ws.max_row]:
        c.font = bold
    for tr in report.trends:
        ws.append([
            tr.operator_code,
            tr.msr_prev if tr.msr_prev is not None else "-",
            tr.msr_this, tr.trend,
        ])

    ws.append([])
    ws.append(["Позначення"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for line in (LEGEND_MSR, LEGEND_MSR_C, LEGEND_CLR, LEGEND_DELTA, LEGEND_SAMPLE):
        ws.append([line])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def monthly_report_to_pdf(report: MonthlyReport) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="Monthly AAR report")
    styles = getSampleStyleSheet()
    story: list = [
        Paragraph(f"Звіт за {report.year}-{report.month:02d}", styles["Title"]),
        Spacer(1, 12),
        Paragraph("Т.4. Інтегральні показники", styles["Heading2"]),
    ]
    integral: list[list[object]] = [[
        "Експл.", "Тип", "Запущ.", "Втрач.", "Ремонт", "Успіх",
        "Успішн., %", "Обслуга, %", "Втрати обсл., %", "Зміна, в.п.",
    ]]
    for r in report.rows:
        integral.append([
            r.operator_code, r.item_type_code, r.launched, r.lost, r.repaired,
            r.success, f"{r.msr:.2%}", f"{r.msr_c:.2%}",
            f"{r.clr:.2%}", f"{r.delta_msr_pp:+.1f}",
        ])
    t = report.totals
    integral.append([
        t.operator_code, t.item_type_code, t.launched, t.lost, t.repaired,
        t.success, f"{t.msr:.2%}", f"{t.msr_c:.2%}",
        f"{t.clr:.2%}", f"{t.delta_msr_pp:+.1f}",
    ])
    story.append(_table(integral))
    story.append(Spacer(1, 6))
    story.extend(
        Paragraph(line, styles["BodyText"])
        for line in (LEGEND_MSR, LEGEND_MSR_C, LEGEND_CLR, LEGEND_DELTA, LEGEND_SAMPLE)
    )

    story.extend([Spacer(1, 12), Paragraph("Рейтинг експлуатантів", styles["Heading2"])])
    rating: list[list[object]] = [
        ["Місце", "Експлуатант", "Успішність обслуги, %", "Запусків", "Категорія"]
    ]
    rating.extend(
        [r.rank, r.operator_code, f"{r.msr_c:.2%}", r.sorties, CATEGORY_UK[r.category]]
        for r in report.rating
    )
    story.append(_table(rating))

    story.extend([Spacer(1, 12), Paragraph("Т.7. Зони відповідальності", styles["Heading2"])])
    zones: list[list[object]] = [["Зона", "Втрати", "Ремонти", "Разом", "Частка"]]
    zones.extend([z.zone.value, z.losses, z.repairs, z.total, f"{z.share_of_launched:.2%}"]
                 for z in report.zones)
    story.append(_table(zones))

    doc.build(story)
    return buf.getvalue()
