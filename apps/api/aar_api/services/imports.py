"""CSV / XLSX bulk import of UsageEvents.

Format (header required, case-insensitive):
  item_serial_no, item_type_code, operator_code, event_date, outcome,
  loss_reason_code, repair_reason_code, notes, client_event_id

Validation per row → if any fails, that row is reported but others still
imported (best-effort, atomic per row). Idempotent via `client_event_id`
when supplied; otherwise duplicates are by-design possible — caller should
provide stable IDs from the source system.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import IO, Any

from openpyxl import load_workbook
from pydantic import ValidationError

from aar_api.models.event import Outcome
from aar_api.schemas.event import UsageEventIn

REQUIRED_HEADERS = {"item_serial_no", "item_type_code", "operator_code", "event_date", "outcome"}
OPTIONAL_HEADERS = {"loss_reason_code", "repair_reason_code", "notes", "client_event_id"}
ALL_HEADERS = REQUIRED_HEADERS | OPTIONAL_HEADERS


@dataclass
class RowError:
    row: int
    message: str


@dataclass
class ImportPreview:
    total_rows: int
    parsed: list[UsageEventIn] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)


def _normalise_headers(headers: list[str]) -> list[str]:
    return [h.strip().lower() for h in headers]


def _check_headers(headers: list[str]) -> str | None:
    norm = set(_normalise_headers(headers))
    missing = REQUIRED_HEADERS - norm
    if missing:
        return f"missing required columns: {sorted(missing)}"
    unknown = norm - ALL_HEADERS
    if unknown:
        return f"unknown columns: {sorted(unknown)}"
    return None


def _parse_date(value: str | date | datetime) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    return date.fromisoformat(s)


def _row_to_dict(headers: list[str], cells: list[Any]) -> dict[str, Any]:
    return {
        h: (str(c).strip() if c is not None and str(c).strip() != "" else None)
        for h, c in zip(headers, cells, strict=False)
    }


def _validate_row(
    row_index: int, raw: dict[str, Any]
) -> tuple[UsageEventIn | None, RowError | None]:
    try:
        event_date = _parse_date(raw["event_date"])
    except (ValueError, TypeError, KeyError) as e:
        return None, RowError(row=row_index, message=f"event_date: {e}")
    try:
        outcome = Outcome(str(raw.get("outcome", "")).lower())
    except ValueError:
        return None, RowError(
            row=row_index,
            message=f"outcome must be one of success/lost/repair, got {raw.get('outcome')!r}",
        )
    try:
        ev = UsageEventIn(
            item_serial_no=str(raw["item_serial_no"]),
            item_type_code=str(raw["item_type_code"]),
            operator_code=str(raw["operator_code"]),
            event_date=event_date,
            outcome=outcome,
            loss_reason_code=raw.get("loss_reason_code"),
            repair_reason_code=raw.get("repair_reason_code"),
            notes=raw.get("notes"),
            client_event_id=raw.get("client_event_id"),
        )
        return ev, None
    except ValidationError as e:
        return None, RowError(row=row_index, message=str(e.errors()[0]["msg"]))
    except KeyError as e:
        return None, RowError(row=row_index, message=f"missing column {e}")


def parse_csv(fp: IO[str]) -> ImportPreview:
    reader = csv.reader(fp)
    try:
        raw_headers = next(reader)
    except StopIteration:
        return ImportPreview(total_rows=0, errors=[RowError(row=0, message="empty file")])
    headers = _normalise_headers(raw_headers)
    err = _check_headers(raw_headers)
    if err:
        return ImportPreview(total_rows=0, errors=[RowError(row=1, message=err)])

    preview = ImportPreview(total_rows=0)
    for i, row in enumerate(reader, start=2):
        preview.total_rows += 1
        raw = _row_to_dict(headers, row)
        ev, row_err = _validate_row(i, raw)
        if ev is not None:
            preview.parsed.append(ev)
        elif row_err is not None:
            preview.errors.append(row_err)
    return preview


def parse_xlsx(data: bytes) -> ImportPreview:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return ImportPreview(total_rows=0, errors=[RowError(row=0, message="empty workbook")])
    rows = ws.iter_rows(values_only=True)
    try:
        raw_headers = list(next(rows))
    except StopIteration:
        return ImportPreview(total_rows=0, errors=[RowError(row=0, message="empty sheet")])
    headers = _normalise_headers([str(h) if h is not None else "" for h in raw_headers])
    err = _check_headers([str(h) if h is not None else "" for h in raw_headers])
    if err:
        return ImportPreview(total_rows=0, errors=[RowError(row=1, message=err)])

    preview = ImportPreview(total_rows=0)
    for i, row in enumerate(rows, start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        preview.total_rows += 1
        raw = _row_to_dict(headers, list(row))
        ev, row_err = _validate_row(i, raw)
        if ev is not None:
            preview.parsed.append(ev)
        elif row_err is not None:
            preview.errors.append(row_err)
    return preview


def parse_bytes(filename: str, data: bytes) -> ImportPreview:
    """Dispatch by extension. .csv → CSV, .xlsx → Excel."""
    name = filename.lower()
    if name.endswith(".csv"):
        return parse_csv(io.StringIO(data.decode("utf-8-sig")))
    if name.endswith((".xlsx", ".xlsm")):
        return parse_xlsx(data)
    return ImportPreview(
        total_rows=0,
        errors=[RowError(row=0, message=f"unsupported file type: {filename}")],
    )
